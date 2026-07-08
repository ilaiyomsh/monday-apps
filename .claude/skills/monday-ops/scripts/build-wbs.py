#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
build-wbs.py — build the optional WBS project-structure board from an Excel file.
3 levels: groups = stages (one Excel sheet per stage), items = milestones (אבן דרך),
subitems = main tasks (משימה עיקרית). Reads the projectStructure board id + column ids
from state.json. Requires `openpyxl`.

Excel format: one sheet per stage (sheet name = stage name). Row 1 = header. Each
subsequent row: col A = milestone (אבן דרך), cols B..D = its main tasks (משימה עיקרית).

config.wbs shape:
{
  "excel": "/abs/path/to/structure.xlsx",
  "stages": ["תכנון מוקדם","תכנון סופי","מכרז","IFC","ליווי ביצוע","סיום פרויקט"],
  "status": { "תכנון מוקדם":"done", "מכרז":"in_progress", "IFC":"not_started" },
  "owner":  { "תכנון מוקדם":48274917, "מכרז":37022703 },
  "linkToPortfolioItem": "12200750452"   // optional: link this WBS board to the project
}

Usage:  ./build-wbs.py <config.json> [--apply] [--confirm-wipe]

DESTRUCTIVE: with --apply this script DELETES EVERY EXISTING ITEM on the
projectStructure board before rebuilding it. It refuses to delete unless you
pass --confirm-wipe or type WIPE at the interactive prompt.
"""
import sys, os, json, subprocess
from pathlib import Path

HERE = os.path.dirname(os.path.abspath(__file__))        # .../monday-ops/scripts
# Canonical API helper (owned by the mapps skill; the local duplicate was removed).
API = os.environ.get(
    "MAPPS_API_SH",
    # Sibling skill: <skills-root>/mapps/mapps-api.sh, derived from this file's
    # own location so the skill works from wherever the repo is cloned.
    str(Path(__file__).resolve().parents[2] / "mapps" / "mapps-api.sh"))
APIV = "2026-04"


def gql(q, v=None, apply=True):
    if q.lstrip().startswith("mutation") and not apply:
        return {"_dryrun": True}
    a = [API, q] + ([json.dumps(v, ensure_ascii=False)] if v is not None else []) + [APIV]
    out = json.loads(subprocess.run(a, capture_output=True, text=True).stdout)
    if out.get("errors"):
        print("  ! error:", json.dumps(out["errors"], ensure_ascii=False)[:200])
    return out.get("data", {}) or {}


def main():
    import openpyxl
    apply = "--apply" in sys.argv
    config = json.load(open(sys.argv[1], encoding="utf-8"))
    state = json.load(open(os.path.join(config["outDir"], "state.json"), encoding="utf-8"))
    wbs = config["wbs"]
    if "projectStructure" not in state["boards"]:
        raise SystemExit("projectStructure board not in state — enable the projectStructure option and run provision.py first.")
    board = state["boards"]["projectStructure"]
    BID = board["id"]
    col = board["columns"]
    lab = board["labels"]

    # read structure from Excel
    wb = openpyxl.load_workbook(wbs["excel"], data_only=True)
    structure = []
    for stage in wbs["stages"]:
        if stage not in wb.sheetnames:
            print(f"  ! sheet '{stage}' missing, skipping"); continue
        ws = wb[stage]
        rows = [[("" if c is None else str(c).strip()) for c in r] for r in ws.iter_rows(values_only=True)]
        milestones = [(r[0], [t for t in r[1:4] if t]) for r in rows[1:] if r and r[0]]
        structure.append((stage, milestones))

    statuses = wbs.get("status", {})
    owners = wbs.get("owner", {})

    # wipe existing items (re-runnable template) — DESTRUCTIVE, gated
    items = gql('query($b:[ID!]){boards(ids:$b){name items_page(limit:200){items{id}}}}', {"b": [BID]}, True)
    board_node = items.get("boards", [{}])[0] if items else {}
    existing_items = board_node.get("items_page", {}).get("items", [])
    if existing_items and apply:
        print("=" * 72)
        print("!! DESTRUCTIVE ACTION !!")
        print(f"!! This will PERMANENTLY DELETE {len(existing_items)} existing item(s)")
        print(f"!! on board '{board_node.get('name', '?')}' (id {BID}) before rebuilding.")
        print("=" * 72)
        if "--confirm-wipe" not in sys.argv:
            if sys.stdin.isatty():
                answer = input("Type WIPE to delete these items, anything else to abort: ")
                if answer.strip() != "WIPE":
                    raise SystemExit("Aborted — no items were deleted.")
            else:
                raise SystemExit(
                    "Refusing to delete existing items without confirmation. "
                    "Re-run with --confirm-wipe (after the user explicitly approved the wipe).")
    for it in existing_items:
        gql('mutation($i:ID!){delete_item(item_id:$i){id}}', {"i": it["id"]}, apply)

    # ensure one group per stage (create missing)
    existing_groups = gql('query($b:[ID!]){boards(ids:$b){groups{id title}}}', {"b": [BID]}, True)
    gmap = {g["title"]: g["id"] for g in (existing_groups.get("boards", [{}])[0].get("groups", []) if existing_groups else [])}
    for stage, _ in structure:
        if stage not in gmap:
            r = gql('mutation($b:ID!,$n:String!){create_group(board_id:$b,group_name:$n){id}}', {"b": BID, "n": stage}, apply)
            gmap[stage] = (r.get("create_group", {}) or {}).get("id", "DRYRUN")

    cm = cs = 0
    for stage, milestones in structure:
        gid = gmap[stage]
        st_key = statuses.get(stage, "not_started")
        st_id = lab.get("status", {}).get(st_key)
        owner = owners.get(stage)
        icv = {}
        if st_id:
            icv[col["status"]] = {"index": int(st_id)}
        if owner:
            icv[col["owner"]] = {"personsAndTeams": [{"id": int(owner), "kind": "person"}]}
        for ms, tasks in milestones:
            r = gql('mutation($b:ID!,$g:String!,$n:String!,$cv:JSON!){create_item(board_id:$b,group_id:$g,item_name:$n,column_values:$cv,create_labels_if_missing:true){id}}',
                    {"b": BID, "g": gid, "n": ms, "cv": json.dumps(icv, ensure_ascii=False)}, apply)
            mid = (r.get("create_item", {}) or {}).get("id")
            if not apply:
                cm += 1; cs += len(tasks); continue
            if not mid:
                continue
            cm += 1
            for t in tasks:
                sr = gql('mutation($p:ID!,$n:String!,$cv:JSON!){create_subitem(parent_item_id:$p,item_name:$n,column_values:$cv,create_labels_if_missing:true){id}}',
                         {"p": mid, "n": t, "cv": json.dumps(icv, ensure_ascii=False)}, apply)
                if (sr.get("create_subitem")):
                    cs += 1

    # optional: link this WBS board to the Portfolio project item
    link_item = wbs.get("linkToPortfolioItem")
    if link_item and "portfolio" in state["boards"] and "link" in state["boards"]["portfolio"]["columns"]:
        pf = state["boards"]["portfolio"]
        # link the FIRST item of the WBS board to the project (board_relation expects item ids)
        first = gql('query($b:[ID!]){boards(ids:$b){items_page(limit:1){items{id}}}}', {"b": [BID]}, True)
        wbs_items = first.get("boards", [{}])[0].get("items_page", {}).get("items", []) if first else []
        if wbs_items:
            gql('mutation($b:ID!,$i:ID!,$c:String!,$v:JSON!){change_column_value(board_id:$b,item_id:$i,column_id:$c,value:$v){id}}',
                {"b": pf["id"], "i": link_item, "c": pf["columns"]["link"],
                 "v": json.dumps({"linkedPulseIds": [{"linkedPulseId": int(wbs_items[0]["id"])}]}, ensure_ascii=False)}, apply)

    print(f"WBS: {cm} milestones + {cs} main tasks {'created' if apply else '(planned)'}")


if __name__ == "__main__":
    main()
